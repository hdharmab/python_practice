"""
This program analyses an excel sheet using openpyxl
and returns:
1. The number of products per supplier
2. The total cost of inventpry per supplier
3. The number of products whose inventory < 10
"""
import openpyxl

name_xl_workbook = input("Enter name of your excel workbook with extension\n >")
sheet_name = input("Enter the name of the excel sheet\n >")
if name_xl_workbook == "":
    name_xl_workbook = "inventory.xlsx"
if sheet_name == "":
    sheet_name = "Sheet1"

inventory_file = openpyxl.load_workbook(name_xl_workbook)
product_list = inventory_file[sheet_name]

total_quantity_products_per_supplier = {}
low_inventory = {}
total_inventory_value_per_supplier = {}

for row in range(2, product_list.max_row + 1):
    # create variables
    product_number = product_list.cell(row, 1).value
    product_inventory = product_list.cell(row, 2).value
    product_price = product_list.cell(row, 3).value
    product_supplier = product_list.cell(row, 4).value
    product_cost = product_inventory * product_price
    # grab each cell of column 5 to later add product_cost
    inventory_cost = product_list.cell(row,5)
    # The number of products per supplier
    if product_supplier in total_quantity_products_per_supplier:
        current_value = total_quantity_products_per_supplier.get(product_supplier)
        total_quantity_products_per_supplier[product_supplier] = current_value + 1
    else:
        total_quantity_products_per_supplier[product_supplier] = 1
    # Inventory is < 10
    if product_inventory < 10:
        if product_number in low_inventory:
            current_inventory = low_inventory.get(product_number)
            low_inventory[product_number] = current_inventory + product_inventory
        else:
            low_inventory[product_number] = product_inventory
    # total cost of inventory per supplier
    if product_supplier in total_inventory_value_per_supplier:
        current_total_cost = total_inventory_value_per_supplier.get(product_supplier)
        total_inventory_value_per_supplier[product_supplier] = current_total_cost + product_cost
    else:
        total_inventory_value_per_supplier[product_supplier] = product_cost
    # add a column called inventory cost to Excel sheet
    inventory_cost.value = product_cost


print(total_quantity_products_per_supplier)
print(total_inventory_value_per_supplier)
print(low_inventory)

column5_header = product_list.cell(1,5)
column5_header.value = "Cost of Inventory"
# save our file as a new file inventory with total value and extension .xlsx
inventory_file.save("my_workbook.xlsx")



